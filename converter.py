from __future__ import annotations

import argparse
import contextlib
import copy
import io
import re
import sys
import tempfile
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
VENDOR_DIR = BASE_DIR / "vendor"
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))

INPUT_NAME = "amsexport.xls"
DEFAULT_OUTPUT_NAME = "amsexport_formatiert.xlsx"
DEFAULT_LOGO_PATH = BASE_DIR / "assets" / "liga-logo.png"

OUTPUT_HEADERS = [
    "Lfd. Nr.",
    "Versicherungs-Nr.",
    "Kennzeichen",
    "Haftpflicht",
    "Schutzbrief",
    "Vollkasko SB",
    "Teilkasko SB",
    "Bruttobeitrag",
    "Fahrgestell-Nr.",
    "Anmerkung",
]


FIELD_PATTERNS = {
    "lfd": ["lfd", "laufende"],
    "policy": ["versicherung", "vertrags", "police", "vsnr", "vnr"],
    "plate": ["kennzeichen", "kfz-kennzeichen", "amtliches"],
    "liability": ["haftpflicht", "hp"],
    "schutzbrief": ["schutzbrief"],
    "vollkasko": ["vollkasko", "vk"],
    "teilkasko": ["teilkasko", "tk"],
    "premium": ["bruttobeitrag", "beitrag", "jahresbeitrag", "praemie", "prämie"],
    "vin": ["fahrgestell", "fahrzeugident", "fin", "vin"],
    "note": ["anmerkung", "bemerkung", "hinweis", "notiz"],
    "customer_name": ["kundenname", "kunde", "versicherungsnehmer"],
    "customer_street": ["strasse", "straße"],
    "customer_zip": ["plz", "postleitzahl"],
    "customer_city": ["ort", "stadt"],
}


@dataclass
class SourceTable:
    headers: list[str]
    rows: list[list[Any]]


@dataclass
class CustomerInfo:
    name: str = ""
    street: str = ""
    zip_code: str = ""
    city: str = ""


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    return re.sub(r"\s+", " ", text)


def convert_xls_to_xlsx_with_excel(input_path: Path, temp_dir: Path) -> Path:
    """Convert old .xls files through locally installed Excel via pywin32."""
    try:
        import win32com.client as win32
    except ImportError as exc:
        raise RuntimeError(
            "Diese .xls-Datei braucht Microsoft Excel/pywin32 zum Einlesen. "
            "pywin32 ist in dieser Python-Umgebung nicht installiert."
        ) from exc

    output_path = temp_dir / f"{input_path.stem}_converted.xlsx"
    excel = None
    workbook = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.EnableEvents = False
        excel.AutomationSecurity = 3

        workbook = excel.Workbooks.Open(
            str(input_path),
            UpdateLinks=0,
            ReadOnly=True,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        workbook.SaveAs(str(output_path), FileFormat=51)
        return output_path
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()


def load_source_workbook(input_path: Path) -> Workbook:
    if input_path.suffix.lower() == ".xlsx":
        return load_workbook(input_path, data_only=True)

    if input_path.suffix.lower() != ".xls":
        raise ValueError(f"Nicht unterstuetztes Dateiformat: {input_path.suffix}")

    with tempfile.TemporaryDirectory(prefix="amsexport_") as temp:
        converted = convert_xls_to_xlsx_with_excel(input_path, Path(temp))
        return load_workbook(converted, data_only=True)


def extract_source_table(input_path: Path) -> SourceTable:
    raw_rows = read_source_rows(input_path)
    raw_rows = [trim_trailing_empty(row) for row in raw_rows]
    raw_rows = [row for row in raw_rows if any(value not in (None, "") for value in row)]

    if not raw_rows:
        raise ValueError("Die Quelldatei enthaelt keine erkennbaren Daten.")

    header_index = find_header_row(raw_rows)
    headers = [str(value).strip() if value is not None else "" for value in raw_rows[header_index]]
    rows = raw_rows[header_index + 1 :]
    return SourceTable(headers=headers, rows=rows)


def read_source_rows(input_path: Path) -> list[list[Any]]:
    if input_path.suffix.lower() == ".xls":
        try:
            return read_xls_rows(input_path)
        except ImportError:
            wb = load_source_workbook(input_path)
            ws = wb.worksheets[0]
            return [
                [cell.value for cell in row]
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)
            ]

    wb = load_source_workbook(input_path)
    ws = wb.worksheets[0]
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column)
    ]


def read_xls_rows(input_path: Path) -> list[list[Any]]:
    import xlrd
    from xlrd.xldate import xldate_as_datetime

    with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
        book = xlrd.open_workbook(str(input_path), logfile=io.StringIO())
    sheet = book.sheet_by_index(0)
    rows: list[list[Any]] = []

    for row_index in range(sheet.nrows):
        row_values: list[Any] = []
        for col_index in range(sheet.ncols):
            cell = sheet.cell(row_index, col_index)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_EMPTY:
                value = None
            elif cell.ctype == xlrd.XL_CELL_DATE:
                value = xldate_as_datetime(value, book.datemode)
            elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                value = int(value)
            row_values.append(value)
        rows.append(row_values)

    return rows


def trim_trailing_empty(row: list[Any]) -> list[Any]:
    result = list(row)
    while result and result[-1] in (None, ""):
        result.pop()
    return result


def find_header_row(rows: list[list[Any]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:30]):
        normalized = " ".join(normalize_text(value) for value in row)
        score = sum(
            1
            for patterns in FIELD_PATTERNS.values()
            if any(pattern in normalized for pattern in patterns)
        )
        filled_cells = sum(1 for value in row if value not in (None, ""))
        score = score * 10 + min(filled_cells, 9)
        if score > best_score:
            best_score = score
            best_index = index
    return best_index


def build_column_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized_headers = [normalize_text(header) for header in headers]

    for field, patterns in FIELD_PATTERNS.items():
        for index, header in enumerate(normalized_headers):
            if any(pattern in header for pattern in patterns):
                mapping[field] = index
                break

    return mapping


def get_mapped_value(row: list[Any], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def first_text_value(table: SourceTable, mapping: dict[str, int], field: str) -> str:
    for row in table.rows:
        value = get_mapped_value(row, mapping, field)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def infer_customer_info(table: SourceTable) -> CustomerInfo:
    mapping = build_column_map(table.headers)
    return CustomerInfo(
        name=first_text_value(table, mapping, "customer_name"),
        street=first_text_value(table, mapping, "customer_street"),
        zip_code=first_text_value(table, mapping, "customer_zip"),
        city=first_text_value(table, mapping, "customer_city"),
    )


def parse_money(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None
    text = text.replace("€", "").replace("EUR", "").replace("\u00a0", " ").strip()
    text = text.replace(".", "").replace(",", ".")
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in ("", "-", "."):
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_deductible(value: Any) -> Any:
    money = parse_money(value)
    if money is not None:
        return money
    return value


def build_output_rows(table: SourceTable) -> list[list[Any]]:
    mapping = build_column_map(table.headers)
    output_rows: list[list[Any]] = []

    for row in table.rows:
        if not any(value not in (None, "") for value in row):
            continue

        lfd = get_mapped_value(row, mapping, "lfd") or len(output_rows) + 1
        output_rows.append(
            [
                lfd,
                get_mapped_value(row, mapping, "policy"),
                get_mapped_value(row, mapping, "plate"),
                get_mapped_value(row, mapping, "liability"),
                get_mapped_value(row, mapping, "schutzbrief"),
                normalize_deductible(get_mapped_value(row, mapping, "vollkasko")),
                normalize_deductible(get_mapped_value(row, mapping, "teilkasko")),
                parse_money(get_mapped_value(row, mapping, "premium")),
                get_mapped_value(row, mapping, "vin"),
                get_mapped_value(row, mapping, "note"),
            ]
        )

    return output_rows


def create_output_workbook(
    output_path: Path,
    output_rows: list[list[Any]],
    stand: str,
    kunde: str,
    adresse: str,
    ort: str,
    logo_path: Path | None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Beitragsuebersicht"

    ws["A1"] = "Beitragsuebersicht KFZ-Versicherung"
    ws["A2"] = f"Stand {stand}"
    ws["A4"] = "Versicherungsnehmer"
    ws["A5"] = kunde
    ws["A6"] = adresse
    ws["A7"] = ort

    ws["I1"] = "LIGA Löffler"
    ws["I2"] = "Versicherungsmakler"
    if logo_path is not None and logo_path.exists():
        ws["I1"] = None
        ws["I2"] = None

    table_start_row = 9
    for col_index, header in enumerate(OUTPUT_HEADERS, start=1):
        ws.cell(row=table_start_row, column=col_index, value=header)

    for row_index, row_values in enumerate(output_rows, start=table_start_row + 1):
        for col_index, value in enumerate(row_values, start=1):
            ws.cell(row=row_index, column=col_index, value=value)

    apply_layout(ws, table_start_row, len(output_rows))
    add_logo(ws, logo_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def add_logo(ws, logo_path: Path | None) -> None:
    if logo_path is None or not logo_path.exists():
        return

    logo = ExcelImage(str(logo_path))
    logo.width = 240
    logo.height = 106
    ws.add_image(logo, "H1")


def apply_layout(ws, table_start_row: int, row_count: int) -> None:
    blue = "005CA8"
    gray = "D9D9D9"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9D9D9")

    ws["A1"].font = Font(name="Arial", size=18, bold=True)
    ws["A2"].font = Font(name="Arial", size=14, bold=True)
    ws["A4"].font = Font(name="Arial", size=11, bold=True)
    for cell in ("A5", "A6", "A7"):
        ws[cell].font = Font(name="Arial", size=11, bold=True)
    ws["I1"].font = Font(name="Arial", size=20, bold=True, color=blue)
    ws["I2"].font = Font(name="Arial", size=10, bold=True, color="777777")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(OUTPUT_HEADERS)):
        for cell in row:
            font = copy.copy(cell.font)
            font.name = "Arial"
            cell.font = font
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = Border(bottom=thin_gray)

    for cell in ws[table_start_row]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(name="Arial", bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first_data_row = table_start_row + 1
    last_data_row = table_start_row + max(row_count, 1)
    for row_number in range(first_data_row, last_data_row + 1):
        fill = PatternFill("solid", fgColor=gray if row_number % 2 == 0 else white)
        for col_number in range(1, len(OUTPUT_HEADERS) + 1):
            ws.cell(row=row_number, column=col_number).fill = fill

    widths = [11, 20, 17, 16, 14, 15, 15, 16, 28, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 16
    for row_number in range(4, 8):
        ws.row_dimensions[row_number].height = 22
    ws.row_dimensions[8].height = 12
    ws.row_dimensions[table_start_row].height = 30
    for row_number in range(first_data_row, last_data_row + 1):
        ws.row_dimensions[row_number].height = 25
    ws.freeze_panes = f"A{table_start_row + 1}"
    ws.auto_filter.ref = f"A{table_start_row}:J{last_data_row}"
    ws.print_title_rows = f"${table_start_row}:${table_start_row}"
    ws.print_area = f"A1:J{last_data_row}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True

    for row_number in range(first_data_row, last_data_row + 1):
        ws.cell(row=row_number, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_number, column=8).number_format = '#,##0.00 "€"'
        ws.cell(row=row_number, column=6).number_format = '#,##0 "€"'
        ws.cell(row=row_number, column=7).number_format = '#,##0 "€"'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Formatiert amsexport.xls als normale, sortierbare KFZ-Beitragsuebersicht."
    )
    parser.add_argument(
        "ordner",
        nargs="?",
        default=".",
        help="Ordner mit amsexport.xls. Standard: aktueller Ordner.",
    )
    parser.add_argument("--input", default=INPUT_NAME, help="Name der Quelldatei im Ordner.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT_NAME, help="Name der Zieldatei.")
    parser.add_argument("--stand", default=datetime.now().strftime("%m.%Y"), help="Stand, z. B. 07.2026.")
    parser.add_argument("--kunde", default="", help="Name des Versicherungsnehmers.")
    parser.add_argument("--adresse", default="", help="Adresse des Versicherungsnehmers.")
    parser.add_argument("--ort", default="", help="PLZ und Ort des Versicherungsnehmers.")
    parser.add_argument(
        "--logo",
        default=str(DEFAULT_LOGO_PATH),
        help="Pfad zum Logo. Standard: assets\\liga-logo.png",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    folder = Path(args.ordner).expanduser().resolve()
    input_path = folder / args.input
    output_path = folder / args.output
    logo_path = Path(args.logo).expanduser().resolve() if args.logo else None

    if not input_path.exists():
        print(f"Quelldatei nicht gefunden: {input_path}", file=sys.stderr)
        return 1

    try:
        source_table = extract_source_table(input_path)
        customer_info = infer_customer_info(source_table)
        output_rows = build_output_rows(source_table)
        create_output_workbook(
            output_path=output_path,
            output_rows=output_rows,
            stand=args.stand,
            kunde=args.kunde or customer_info.name,
            adresse=args.adresse or customer_info.street,
            ort=args.ort or " ".join(
                part for part in (customer_info.zip_code, customer_info.city) if part
            ),
            logo_path=logo_path,
        )
    except Exception as exc:
        print(f"Fehler: {exc}", file=sys.stderr)
        return 1

    print(f"Fertig: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
